import os
import hashlib
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook

class App:
    def __init__(self, master):
        self.master = master
        master.title("文件MD5哈希值计算器")
        #master.iconbitmap("favicon.ico")

        # 菜单栏
        self.menu = tk.Menu(master)
        self.file_menu = tk.Menu(self.menu, tearoff=0)
        self.file_menu.add_command(label="退出", command=master.quit)
        self.menu.add_cascade(label="文件", menu=self.file_menu)
        master.config(menu=self.menu)

        # 添加选择文件夹按钮
        self.folder_label = tk.Label(master, text="选择文件夹：")
        self.folder_label.grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.folder_button = ttk.Button(master, text="选择", command=self.choose_folder)
        self.folder_button.grid(row=0, column=1, padx=5, pady=5, sticky='w')

        # 添加选择保存位置按钮
        self.save_label = tk.Label(master, text="保存结果到：")
        self.save_label.grid(row=1, column=0, padx=5, pady=5, sticky='w')
        self.save_button = ttk.Button(master, text="选择", command=self.choose_save_path)
        self.save_button.grid(row=1, column=1, padx=5, pady=5, sticky='w')

        # 添加计算按钮
        self.calculate_button = ttk.Button(master, text="计算MD5哈希值", command=self.calculate)
        self.calculate_button.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

        # 添加进度条
        self.progressbar = ttk.Progressbar(master, mode='indeterminate')
        self.progressbar.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky='w')

    # 选择文件夹
    def choose_folder(self):
        self.folder_path = filedialog.askdirectory()
        self.folder_button.configure(text=self.folder_path)

    # 选择保存位置
    def choose_save_path(self):
        self.save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')])
        self.save_button.configure(text=self.save_path)

    # 计算MD5哈希值
    def calculate(self):
        if not hasattr(self, 'folder_path') or not hasattr(self, 'save_path'):
            messagebox.showerror("错误", "请选择文件夹和保存位置。")
            return

        self.progressbar.start()
        files = []
        for dirpath, dirnames, filenames in os.walk(self.folder_path):
            for filename in filenames:
                filepath = os.path.join(dirpath, filename)
                with open(filepath, "rb") as f:
                    filehash = hashlib.md5()
                    while chunk := f.read(4096):
                        filehash.update(chunk)
                    files.append((filename, filehash.hexdigest()))

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='文件名')
        ws.cell(row=1, column=2, value='MD5哈希值')
        for row in range(len(files)):
            ws.cell(row=row + 2, column=1, value=files[row][0])
            ws.cell(row=row + 2, column=2, value=files[row][1])
        wb.save(self.save_path)
        messagebox.showinfo("成功", "MD5哈希值计算完成，结果已保存。")
        self.progressbar.stop()

# 启动程序
root = tk.Tk()
app = App(root)
root.mainloop()
